#!/usr/bin/env python3
"""
整合的Excel数据预览工具
自动生成包含所有预览sheets的Excel文件

使用方法:
  python3 import_all_preview.py --year 2024

生成的Excel文件包含4个sheets:
  1. 2024-expense-USD: 美国总账的费用预览
  2. 2024-expense-CNY: 中国总账的费用预览
  3. 2024-budgets-USD: 美国总账的预算预览
  4. 2024-budgets-CNY: 中国总账的预算预览

参数说明:
  --year: 年份 (必填)
  --family: 家庭ID (默认: 1)
  --mapping: 分类映射文件 (默认: category_mapping_corrected.json)
  --output: 输出文件名 (默认: preview_{year}.xlsx)
"""

import pandas as pd
import json
import sys
import argparse
import os
import subprocess
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def run_preview_command(command, description):
    """运行预览命令并返回结果"""
    print(f"\n{'='*80}")
    print(f"🔄 {description}")
    print(f"{'='*80}")
    print(f"命令: {' '.join(command)}")

    try:
        result = subprocess.run(command, capture_output=True, text=True, check=True)
        print(result.stdout)
        if result.stderr:
            print(result.stderr)
        return True
    except subprocess.CalledProcessError as e:
        print(f"❌ 错误: {e}")
        print(f"输出: {e.stdout}")
        print(f"错误: {e.stderr}")
        return False

def load_preview_json(json_file):
    """加载JSON预览文件并转换为DataFrame"""
    if not os.path.exists(json_file):
        print(f"⚠️  警告: JSON文件不存在: {json_file}")
        return None

    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # 提取records部分
        records = data.get('records', [])
        if not records:
            print(f"⚠️  警告: JSON文件中没有记录: {json_file}")
            return None

        # 转换为DataFrame
        df = pd.DataFrame(records)
        return df
    except Exception as e:
        print(f"❌ 错误: 无法读取JSON文件 {json_file}: {e}")
        return None

def format_worksheet(ws, df, title):
    """格式化工作表样式"""
    # 标题行样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 设置标题行格式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 设置数据行格式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 最大50字符宽度
        ws.column_dimensions[column_letter].width = adjusted_width

    # 冻结首行
    ws.freeze_panes = ws['A2']

def create_preview_excel(year, family_id, mapping_file, output_file):
    """创建包含所有预览sheets的Excel文件"""

    print(f"\n{'='*80}")
    print(f"📊 开始生成 {year} 年度预览文件")
    print(f"{'='*80}")
    print(f"年份: {year}")
    print(f"家庭ID: {family_id}")
    print(f"映射文件: {mapping_file}")
    print(f"输出文件: {output_file}")
    print()

    # 脚本在scripts/目录下，需要找到import/目录
    script_dir = os.path.dirname(os.path.abspath(__file__))  # scripts/目录
    import_dir = os.path.dirname(script_dir)  # import/目录

    # 检查输入文件
    excel_file = f"{year}.xlsx"
    excel_path = os.path.join(import_dir, excel_file)
    if not os.path.exists(excel_path):
        print(f"❌ 错误: Excel文件不存在: {excel_path}")
        sys.exit(1)

    mapping_path = os.path.join(import_dir, mapping_file)
    if not os.path.exists(mapping_path):
        print(f"❌ 错误: 映射文件不存在: {mapping_path}")
        sys.exit(1)

    # 定义所有预览任务
    tasks = [
        {
            'name': f'{year}-expense-USD',
            'description': '美国总账 - 费用预览',
            'script': 'import_expenses.py',
            'json_file': os.path.join(import_dir, f'preview_{year}_USD.json'),
            'command': [
                'python3', os.path.join(script_dir, 'import_expenses.py'), 'preview',
                '--file', excel_path,
                '--sheet', f'{year}总帐 (US)',
                '--family', str(family_id),
                '--year', str(year),
                '--currency', 'USD',
                '--mapping', mapping_path
            ]
        },
        {
            'name': f'{year}-expense-CNY',
            'description': '中国总账 - 费用预览',
            'script': 'import_expenses.py',
            'json_file': os.path.join(import_dir, f'preview_{year}_CNY.json'),
            'command': [
                'python3', os.path.join(script_dir, 'import_expenses.py'), 'preview',
                '--file', excel_path,
                '--sheet', f'{year}总帐（中国）',
                '--family', str(family_id),
                '--year', str(year),
                '--currency', 'CNY',
                '--mapping', mapping_path
            ]
        },
        {
            'name': f'{year}-budgets-USD',
            'description': '美国总账 - 预算预览',
            'script': 'import_budgets.py',
            'json_file': os.path.join(import_dir, f'budget_preview_{year}_USD.json'),
            'command': [
                'python3', os.path.join(script_dir, 'import_budgets.py'), 'preview',
                '--file', excel_path,
                '--sheet', f'{year}总帐 (US)',
                '--family', str(family_id),
                '--year', str(year),
                '--currency', 'USD',
                '--mapping', mapping_path
            ]
        },
        {
            'name': f'{year}-budgets-CNY',
            'description': '中国总账 - 预算预览',
            'script': 'import_budgets.py',
            'json_file': os.path.join(import_dir, f'budget_preview_{year}_CNY.json'),
            'command': [
                'python3', os.path.join(script_dir, 'import_budgets.py'), 'preview',
                '--file', excel_path,
                '--sheet', f'{year}总帐（中国）',
                '--family', str(family_id),
                '--year', str(year),
                '--currency', 'CNY',
                '--mapping', mapping_path
            ]
        }
    ]

    # 执行所有预览命令
    print(f"\n{'='*80}")
    print("第1步: 生成所有预览JSON文件")
    print(f"{'='*80}")

    successful_tasks = []
    for task in tasks:
        success = run_preview_command(task['command'], task['description'])
        if success:
            successful_tasks.append(task)
        else:
            print(f"⚠️  跳过sheet: {task['name']}")

    if not successful_tasks:
        print("\n❌ 错误: 没有成功生成任何预览数据")
        sys.exit(1)

    # 创建Excel工作簿
    print(f"\n{'='*80}")
    print("第2步: 创建预览Excel文件")
    print(f"{'='*80}")

    wb = Workbook()
    wb.remove(wb.active)  # 移除默认sheet

    sheets_created = 0
    for task in successful_tasks:
        print(f"\n📄 处理sheet: {task['name']}")

        # 加载JSON数据
        df = load_preview_json(task['json_file'])
        if df is None:
            print(f"⚠️  跳过sheet: {task['name']}")
            continue

        # 创建新sheet
        ws = wb.create_sheet(title=task['name'])

        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

        # 格式化
        format_worksheet(ws, df, task['name'])

        print(f"  ✅ Sheet创建成功: {len(df)} 条记录")
        sheets_created += 1

    if sheets_created == 0:
        print("\n❌ 错误: 没有创建任何sheet")
        sys.exit(1)

    # 保存文件
    try:
        # 确保output_file使用绝对路径（指向import目录）
        if not os.path.isabs(output_file):
            output_file = os.path.join(import_dir, output_file)
        wb.save(output_file)
        print(f"\n{'='*80}")
        print(f"✅ 预览文件生成成功!")
        print(f"{'='*80}")
        print(f"📁 文件路径: {os.path.abspath(output_file)}")
        print(f"📊 包含sheets: {sheets_created}")
        print()

        # 显示每个sheet的统计
        print("Sheet详情:")
        for task in successful_tasks:
            df = load_preview_json(task['json_file'])
            if df is not None:
                print(f"  • {task['name']}: {len(df)} 条记录")

        # 清理中间JSON文件
        print(f"\n🧹 清理中间文件...")
        cleaned_count = 0
        for task in successful_tasks:
            if os.path.exists(task['json_file']):
                try:
                    os.remove(task['json_file'])
                    print(f"  ✓ 已删除: {task['json_file']}")
                    cleaned_count += 1
                except Exception as e:
                    print(f"  ⚠️  无法删除 {task['json_file']}: {e}")

        # 删除单独生成的preview xlsx文件
        for task in successful_tasks:
            preview_xlsx = task['json_file'].replace('.json', '.xlsx')
            if os.path.exists(preview_xlsx):
                try:
                    os.remove(preview_xlsx)
                    print(f"  ✓ 已删除: {preview_xlsx}")
                    cleaned_count += 1
                except Exception as e:
                    print(f"  ⚠️  无法删除 {preview_xlsx}: {e}")

        if cleaned_count > 0:
            print(f"  共清理 {cleaned_count} 个中间文件")

        print(f"\n💡 下一步:")
        print(f"  1. 打开 {output_file} 检查预览数据")
        print(f"  2. 确认数据无误即可")
        print(f"\n注意: 如需导入数据，请重新运行对应的preview命令生成JSON文件")

    except Exception as e:
        print(f"\n❌ 错误: 无法保存Excel文件: {e}")
        sys.exit(1)

def main():
    parser = argparse.ArgumentParser(
        description='整合的Excel数据预览工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 生成2024年的预览文件
  python3 import_all_preview.py --year 2024

  # 指定输出文件名
  python3 import_all_preview.py --year 2024 --output my_preview.xlsx
        """
    )

    parser.add_argument('--year', type=int, required=True,
                        help='年份 (例如: 2024)')
    parser.add_argument('--family', type=int, default=1,
                        help='家庭ID (默认: 1)')
    parser.add_argument('--mapping', type=str, default='category_mapping_corrected.json',
                        help='分类映射文件 (默认: category_mapping_corrected.json)')
    parser.add_argument('--output', type=str, default=None,
                        help='输出文件名 (默认: preview_{year}.xlsx)')

    args = parser.parse_args()

    # 设置输出文件名
    if args.output is None:
        args.output = f'preview_{args.year}.xlsx'

    # 创建预览文件
    create_preview_excel(args.year, args.family, args.mapping, args.output)

if __name__ == '__main__':
    main()
